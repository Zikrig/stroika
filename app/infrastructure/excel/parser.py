from io import BytesIO

from openpyxl import load_workbook

from app.domain.rules import validate_pdo_formula


EXPECTED_HEADERS = [
    "ID заявки",
    "С кем согласовано",
    "Дата/время заявки",
    "Дата потребности",
    "Подобъект",
    "Прораб",
    "Наименование от прораба",
    "Наименование по 1С",
    "Код 1С",
    "Запрошено",
    "Ед. изм.",
    "Со склада",
    "В закупку",
]


def _parse_sheet_rows(rows: list[tuple]) -> list[dict]:
    """Из уже прочитанных строк листа (первая — заголовок) собирает список словарей."""
    if len(rows) < 2:
        return []
    header = [str(v).strip() if v is not None else "" for v in rows[0]]
    if header[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        return []
    result: list[dict] = []
    for line in rows[1:]:
        if line[0] is None:
            continue
        requested = float(line[9] or 0)
        from_stock = float(line[11] or 0)
        to_purchase = float(line[12] or 0)
        validate_pdo_formula(requested, from_stock, to_purchase)
        result.append(
            {
                "request_code": str(line[0]).strip(),
                "approved_by": str(line[1] or "").strip(),
                "request_datetime": str(line[2] or "").strip(),
                "need_by": str(line[3] or "").strip(),
                "subobject_name": str(line[4] or "").strip(),
                "foreman_name": str(line[5] or "").strip(),
                "name_from_foreman": str(line[6] or "").strip(),
                "nomenclature_1c": str(line[7] or "").strip(),
                "code_1c": str(line[8] or "").strip(),
                "requested_qty": requested,
                "unit": str(line[10] or "шт").strip(),
                "from_stock_qty": from_stock,
                "to_purchase_qty": to_purchase,
            }
        )
    return result


def parse_pdo_excel(raw: bytes) -> list[dict]:
    """Читает все листы книги, находит листы с ожидаемым заголовком формы ПДО и объединяет строки.
    Листы с названием, содержащим .0 (например шаблон по заявке), не пропускаются."""
    wb = load_workbook(filename=BytesIO(raw), data_only=True)
    result: list[dict] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        result.extend(_parse_sheet_rows(rows))
    if not result:
        raise ValueError("Нет строк с данными или неверные колонки Excel-формы")
    return result
